import pandas as pd
from openpyxl import Workbook

def main(uploaded_file, novo_caminho):
    # Caminho do arquivo Excel
    caminho = uploaded_file
    tabela = pd.read_excel(caminho)

    # URL do arquivo Excel no GitHub (caminho local)
    url = r"C:\Users\henrique.canhadas\OneDrive - Servmar Ambientais\Documentos\Codigos\GitHub Tecnologia\ProjetoCanhadas\Tabelas Consulta\Banco de Dados\banco de dados - VAPOR SOLUTIONS.xlsx"
    tabela_bd = pd.read_excel(url)

    colunas_merge = ["Análise"]
    tabela_merge = pd.merge(tabela, tabela_bd, how='left', on=colunas_merge)

    # Formata a coluna Data da Coleta para o padrão desejado
    tabela_merge['Data da Coleta'] = pd.to_datetime(tabela_merge['Data da Coleta'], format='%m/%d/%Y %H:%M').dt.strftime('%d/%m/%Y %H:%M')
    
    # Removendo valores NaN antes de criar a lista
    lista_descricao_metodo = tabela_merge['Método de Análise'].dropna().unique()

    # Cria um novo DataFrame para armazenar os resultados filtrados e cria um arquivo excel no openpyxl
    resultado_final = pd.DataFrame()
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove a aba padrão que é criada ao iniciar um novo Workbook

    # Filtra e cria aba para dados sem descrição de Método de Análise
    sem_metodo = tabela_merge[tabela_merge['Método de Análise'].isna()]
    if not sem_metodo.empty:
        aba_sem_metodo = workbook.create_sheet(title="Sem Método de Análise")
        colunas_desejadas = ["Identificação das Amostras", "Data da Coleta", "Cas Number", "Análise", "Resultado", "Unidade"]
        aba_sem_metodo.append(colunas_desejadas)
        for linha in sem_metodo[colunas_desejadas].itertuples(index=False):
            aba_sem_metodo.append(linha)

    for descricao_metodo in lista_descricao_metodo:
        filtro = (tabela_merge['Método de Análise'] == descricao_metodo)
        tabela_filtrada = tabela_merge[filtro]

        if not tabela_filtrada.empty:
            # Truncar o título se necessário para evitar erro com nomes de abas longos no Excel
            aba_titulo = descricao_metodo[:31] if len(descricao_metodo) > 31 else descricao_metodo
            aba_descricao_metodo = workbook.create_sheet(title=aba_titulo)
            aba_descricao_metodo.append(colunas_desejadas)

            for linha in tabela_filtrada[colunas_desejadas].itertuples(index=False):
                aba_descricao_metodo.append(linha)

    # Alterar cabeçalhos de todas as abas
    new_headers = ["SAMPLENAME", "SAMPDATE", "CASNUMBER_x", "ANALYTE", "Result", "UNITS", "Description"]
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(new_headers)):
            for cell, new_header in zip(row, new_headers):
                cell.value = new_header
                
    # Salva a tabela merge em "Resultado_Final_Com_Abas.xlsx"
    workbook.save(novo_caminho)

if __name__ == "__main__":
    main()
