import openpyxl
import requests
from io import BytesIO

def carregar_planilha(url):
    response = requests.get(url)
    if response.status_code == 200:
        return openpyxl.load_workbook(BytesIO(response.content))
    else:
        raise Exception(f"Falha ao carregar o arquivo de {url}")

def main(novo_caminho,valor_primario, ordem_planilhas,valor_secundario, ordem_planilhas2):

    caminho_resultado = novo_caminho
    caminho_cetesb = 'https://raw.githubusercontent.com/TecnologiaServmar/ProjetoCanhadas/main/Tabelas%20Consulta/Tabelas/Tabela_Cetesb.xlsx'
    caminho_epa = 'https://raw.githubusercontent.com/TecnologiaServmar/ProjetoCanhadas/main/Tabelas%20Consulta/Tabelas/Tabela_EPA.xlsx'
    caminho_listaholandesa = 'https://raw.githubusercontent.com/TecnologiaServmar/ProjetoCanhadas/main/Tabelas%20Consulta/Tabelas/Tabela_ListaHolandesa.xlsx'
    caminho_conama = 'https://raw.githubusercontent.com/TecnologiaServmar/ProjetoCanhadas/main/Tabelas%20Consulta/Tabelas/Tabela_Conama.xlsx'
    caminho_analise = novo_caminho

    # Carregar as planilhas
    wb_resultado = openpyxl.load_workbook(caminho_resultado)
    wb_cetesb = carregar_planilha(caminho_cetesb)
    wb_epa = carregar_planilha(caminho_epa)
    wb_listaholandesa = carregar_planilha(caminho_listaholandesa)
    wb_conama = carregar_planilha(caminho_conama)

    from openpyxl.styles import Font

#------------------------------------------------------------------------------------------------------------------------------------------------

    for sheet_name in wb_resultado.sheetnames:
        sheet_resultado = wb_resultado[sheet_name]

        # Escolher entre wb_cetesb e wb_epa com base na escolha do usuário
        sheet_escolhido = None
        if ordem_planilhas == 'c':
            sheet_escolhido = wb_cetesb['Sheet1']
        elif ordem_planilhas == 'e':
            sheet_escolhido = wb_epa['Sheet1']
        elif ordem_planilhas == 'l':
            sheet_escolhido = wb_listaholandesa['Sheet1']
        elif ordem_planilhas == 'o':
            sheet_escolhido = wb_conama['Sheet1']

        # Variável para armazenar o índice da coluna onde o primeiro dado é inserido
        indice_coluna_padrao = None

        # Variável para rastrear se o último dado foi lido
        ultimo_dado_lido = False

        # Iterar sobre as linhas da coluna 'CAS' em Resultado_Final_Organizado
        for row_resultado in sheet_resultado.iter_rows(min_row=2, max_row=sheet_resultado.max_row, min_col=2,
                                                        max_col=sheet_resultado.max_column):
            # Obter o valor de 'CAS' em Resultado_Final_Organizado
            cas_value = row_resultado[0].value
            unidade = row_resultado[1].value

            # Verificar se o valor de 'CAS' não é nulo
            if cas_value is not None and unidade is not None:
                # Inicializar uma variável para verificar se houve correspondência
                correspondencia_encontrada = False

                # Inicializar uma lista para armazenar todos os valores da linha correspondente
                valores_associados = []

                # Iterar sobre as linhas da coluna 'CAS' em Tabela_Cetesb ou Tabela_EPA
                for row_primairo in sheet_escolhido.iter_rows(min_row=2, max_row=sheet_escolhido.max_row, min_col=2,
                                                          max_col=sheet_escolhido.max_column):
                    # Verificar se encontrou uma correspondência
                    if row_primairo[0].value == cas_value:
                        # Adicionar todos os valores da linha correspondente à lista
                        valores_associados = [cell.value for cell in row_primairo]

                        # Definir que houve uma correspondência
                        correspondencia_encontrada = True

                        # Sair do loop, pois já encontrou uma correspondência
                        break

                # Se houve correspondência, comparar o quinto item da lista com os valores na mesma linha em Resultado_Final_Organizado.xlsx
                if correspondencia_encontrada and len(valores_associados) >= 1:
                    coluna_selecionada = valores_associados[valor_primario]
                    try:
                        if coluna_selecionada != '-' and coluna_selecionada is not None:
                            coluna_selecionada = float(coluna_selecionada)

                            if unidade == 'mg/L':
                                coluna_selecionada /= 1000
                            elif unidade == 'µg/L':
                                pass
                            else:
                                print(unidade)
                        else:
                            coluna_selecionada = None
                            pass
                        
                    except ValueError:
                        # Se a conversão para int ou float falhar, não faz nada
                        pass
                    

                    # Se esta é a primeira iteração, definir a coluna padrão
                    if indice_coluna_padrao is None:
                        indice_coluna_padrao = sheet_resultado.max_column + 1

                    # Comparar e aplicar a formatação desejada
                    for col_index, cell_resultado in enumerate(row_resultado, start=indice_coluna_padrao):
                        valor_na_mesma_linha = cell_resultado.value

                        try:
                            # Verificar se o valor contém "<" e realizar a comparação
                            if valor_na_mesma_linha is not None and coluna_selecionada is not None:
                                if '<' in str(valor_na_mesma_linha):
                                    cell_resultado.font = Font(color="C0C0C0")  # Pintar o texto de cinza
                                elif float(coluna_selecionada) > float(str(valor_na_mesma_linha).replace(',', '.')):
                                    cell_resultado.font = Font(color="000000")  # Pintar o texto de cinza
                                elif float(coluna_selecionada) < float(str(valor_na_mesma_linha).replace(',', '.')):
                                    cell_resultado.font = Font(color="FF0000")  # Pintar o texto de vermelho
                        except ValueError:
                            # Se a conversão falhar, apenas ignore e continue
                            pass

                    # Adicionar o valor da coluna selecionada como uma nova célula na mesma coluna
                    sheet_resultado.cell(row=row_resultado[0].row, column=indice_coluna_padrao, value=coluna_selecionada)

                else:
                    # Se esta é a primeira iteração, definir a coluna padrão
                    if indice_coluna_padrao is None:
                        indice_coluna_padrao = sheet_resultado.max_column + 1
                    # Adicionar o valor da coluna selecionada como uma nova célula na mesma coluna
                    sheet_resultado.cell(row=row_resultado[0].row, column=indice_coluna_padrao, value='n.e')
                        # Se esta é a primeira iteração, definir a coluna padrão

        if indice_coluna_padrao is None:
            indice_coluna_padrao = sheet_resultado.max_column + 1

        # Marcador de último dado lido
        ultimo_dado_lido = True

        # Imprimir "Valor Cetesb" ou "Valor EPA" uma vez acima do primeiro valor após o último dado
        # Imprimir "Valor Cetesb", "Valor EPA", "Valor Lista Holandesa" ou "Valor Conama" uma vez acima do primeiro valor após o último dado
        if ultimo_dado_lido:
            if ordem_planilhas == 'c':
                sheet_resultado.cell(row=2, column=indice_coluna_padrao, value="Valor Cetesb")
            elif ordem_planilhas == 'e':
                sheet_resultado.cell(row=2, column=indice_coluna_padrao, value="Valor EPA")
            elif ordem_planilhas == 'l':
                sheet_resultado.cell(row=2, column=indice_coluna_padrao, value="Valor Lista Holandesa")
            elif ordem_planilhas == 'o':
                sheet_resultado.cell(row=2, column=indice_coluna_padrao, value="Valor Conama")

            # Encontrar a última linha onde a primeira coluna contém um valor diferente de None
            max_row = sheet_resultado.max_row
            
            # Encontrar a última linha onde a primeira coluna contém um valor diferente de None
            for row_resultado in sheet_resultado.iter_rows(min_row=3, max_row=sheet_resultado.max_row,
                                                            min_col=1, max_col=1):  # Itera sobre a coluna A
                if row_resultado[0].value is None:  # Verifica se a célula na coluna A é None
                    max_row = row_resultado[0].row - 1  # Define max_row como a linha anterior à primeira célula None na coluna A
                    break  # Sai do loop após encontrar a primeira célula None na coluna A
                
            # Adicionar "n.e" às células em branco após inserir os valores do último dado lido
            for row_resultado in sheet_resultado.iter_rows(min_row=3, max_row=max_row,
                                                            min_col=indice_coluna_padrao, max_col=sheet_resultado.max_column):
                for cell_resultado in row_resultado:
                    if cell_resultado.value is None:
                        cell_resultado.value = 'n.e'
            
    # Salvar as alterações na planilha Resultado_Final_Organizado.xlsx
    wb_resultado.save(caminho_analise)

    # Carregar novamente a planilha Resultado_Final_Organizado após as alterações
    wb_resultado = openpyxl.load_workbook(caminho_analise)

#------------------------------------------------------------------------------------------------------------------------------------------------

    # Iterar sobre as abas da planilha Resultado_Final_Organizado após as alterações
    for sheet_name in wb_resultado.sheetnames:
        sheet_resultado = wb_resultado[sheet_name]
        # Segundo loop para processar 'c' se 'e' foi escolhido no primeiro loop, e vice-versa
        if ordem_planilhas2 == 'c':
            # Configurar 'c' para o segundo loop
            sheet_escolhido_secundario = wb_cetesb['Sheet1']
        elif ordem_planilhas2 == 'e':
            # Configurar 'e' para o segundo loop
            sheet_escolhido_secundario = wb_epa['Sheet1']
        elif ordem_planilhas2 == 'l':
            # Configurar 'e' para o segundo loop
            sheet_escolhido_secundario = wb_listaholandesa['Sheet1']
        elif ordem_planilhas2 == 'o':
            # Configurar 'e' para o segundo loop
            sheet_escolhido_secundario = wb_conama['Sheet1']

        # Encontrar a coluna "Valor Cetesb" ou "Valor EPA"
        indice_coluna_primaria = None
        for col_index, col in enumerate(sheet_resultado.iter_cols(min_row=2, max_row=2), start=1):
            # Verificar se o valor atual da célula contém "Valor Cetesb" ou "Valor EPA"
            if "Valor Cetesb" in str(col[0].value) or "Valor EPA" in str(col[0].value) or "Valor Lista Holandesa" in str(col[0].value) or "Valor Conama" in str(col[0].value):
                indice_coluna_primaria = col_index
                break

        # Verificar se a coluna "Valor Cetesb" foi encontrada
        # Verificar se a coluna "Valor Cetesb" foi encontrada
        if indice_coluna_primaria is not None:
            # Adicionar uma nova coluna para "Valor EPA" ou "Valor Cetesb" (invertido)
            indice_coluna_secundaria = indice_coluna_primaria + 1

            if ordem_planilhas2 == 'c':
                titulo_coluna_secundaria = "Valor Cetesb"
            elif ordem_planilhas2 == 'e':
                titulo_coluna_secundaria = "Valor EPA"
            elif ordem_planilhas2 == 'l':
                titulo_coluna_secundaria = "Valor Lista Holandesa"
            elif ordem_planilhas2 == 'o':
                titulo_coluna_secundaria = "Valor Conama"

            sheet_resultado.cell(row=2, column=indice_coluna_secundaria, value=titulo_coluna_secundaria)

            # Iterar sobre as linhas a partir da terceira linha (índice 3)
            for row in range(3, sheet_resultado.max_row + 1):
                valor_primario = sheet_resultado.cell(row=row, column=indice_coluna_primaria).value

                # Comparar e inserir valores na coluna "Valor Secundario"
                if valor_primario in ["n.e"]:

                    cas_value = sheet_resultado.cell(row=row, column=2).value  # Valor da coluna 'CAS'
                    unidade = sheet_resultado.cell(row=row, column=3).value

                    # Verificar se o valor de 'CAS' não é nulo
                    if cas_value is not None and unidade is not None:
                        valores_associados = []  # Inicializar uma lista para armazenar todos os valores da linha correspondente

                        # Iterar sobre as linhas da coluna 'CAS' na planilha secundária (wb_cetesb ou wb_epa)
                        for row_secundaria in sheet_escolhido_secundario.iter_rows(min_row=2, max_row=sheet_escolhido_secundario.max_row, min_col=2, max_col=sheet_escolhido_secundario.max_column):
                            if row_secundaria[0].value == cas_value:
                                valores_associados = [cell.value for cell in row_secundaria]
                                break

                        # Se houve correspondência, comparar o quinto item da lista com os valores na mesma linha em Resultado_Final_Organizado.xlsx
                        # Comparar e inserir valores na coluna "Valor Secundario"
                        if len(valores_associados) >= 1:
                            coluna_selecionada = valores_associados[valor_secundario]

                            try:
                                if coluna_selecionada != '-' and coluna_selecionada is not None:
                                    coluna_selecionada = float(coluna_selecionada)

                                    if unidade == 'mg/L':
                                        coluna_selecionada /= 1000
                                    elif unidade == 'µg/L':
                                        pass
                                    else:
                                        print(unidade)
                                else:
                                    coluna_selecionada = None
                                    pass
                                
                            except ValueError:
                                # Se a conversão para int ou float falhar, não faz nada
                                pass

                            # Se esta é a primeira iteração, definir a coluna padrão
                            if indice_coluna_padrao is None:
                                indice_coluna_padrao = sheet_resultado.max_column + 1

                            # Comparar e aplicar a formatação desejada
                            for col_index in range(2, col_index + 1):
                                cell_resultado = sheet_resultado.cell(row=row, column=col_index)
                                valor_na_mesma_linha = cell_resultado.value

                                try:
                                    # Verificar se a célula não é nula e se o valor contém "<" e realizar a comparação
                                    if valor_na_mesma_linha is not None and coluna_selecionada is not None:
                                        if '<' in str(valor_na_mesma_linha):
                                            cell_resultado.font = Font(color="C0C0C0")  # Pintar o texto de cinza
                                        elif float(coluna_selecionada) > float(str(valor_na_mesma_linha).replace(',', '.')):
                                            cell_resultado.font = Font(color="000000")  # Pintar o texto de cinza
                                        elif float(coluna_selecionada) < float(str(valor_na_mesma_linha).replace(',', '.')):
                                            cell_resultado.font = Font(color="FF0000")  # Pintar o texto de vermelho
                                except ValueError:
                                    # Se a conversão falhar, apenas ignore e continue
                                    pass

                            # Adicionar o valor da coluna selecionada como uma nova célula na mesma coluna
                            sheet_resultado.cell(row=row, column=indice_coluna_secundaria, value=coluna_selecionada)

                            # Marcador de último dado lido
                            ultimo_dado_lido = True
                else:
                    sheet_resultado.cell(row=row, column=indice_coluna_secundaria, value="n.a")

        # Imprimir "Valor Cetesb" ou "Valor EPA" uma vez acima do primeiro valor após o último dado
        if ultimo_dado_lido:
            # Adicionar "n.e" às células em branco após inserir os valores do último dado lido
            for row_resultado in sheet_resultado.iter_rows(min_row=3, max_row=sheet_resultado.max_row,
                                                            min_col=indice_coluna_padrao, max_col=sheet_resultado.max_column):
                for cell_resultado in row_resultado:
                    if cell_resultado.value is None:
                        cell_resultado.value = 'n.e'

    # Salvar as alterações na planilha Resultado_Final_Organizado_Formatado_cetesb_epa.xlsx
    wb_resultado.save(caminho_analise)

    # Carregar novamente a planilha Resultado_Final_Organizado após as alterações
    wb_resultado = openpyxl.load_workbook(caminho_analise)

    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side

    # Carregar novamente a planilha Resultado_Final_Organizado após as alterações
    wb_resultado = openpyxl.load_workbook(caminho_analise)

    # Iterar sobre as abas da planilha Resultado_Final_Organizado após as alterações
    for sheet_name in wb_resultado.sheetnames:
        sheet_resultado = wb_resultado[sheet_name]

        # Inserir uma nova coluna vazia na coluna D
        sheet_resultado.insert_cols(idx=4)
        # Inserir uma nova coluna vazia na coluna E
        sheet_resultado.insert_cols(idx=5)

        # Inicializar variáveis para armazenar as colunas de "Valor Cetesb" e "Valor EPA"
        coluna_cetesb = None
        coluna_epa = None
        coluna_listaholandesa = None
        coluna_conama = None

        # Iterar sobre as células da segunda linha
        for col_idx in range(1, sheet_resultado.max_column + 1):
            valor = sheet_resultado.cell(row=2, column=col_idx).value

            if valor == "Valor Cetesb":
                coluna_cetesb = col_idx
            elif valor == "Valor EPA":
                coluna_epa = col_idx
            elif valor == "Valor Lista Holandesa":
                coluna_listaholandesa = col_idx
            elif valor == "Valor Conama":
                coluna_conama = col_idx

        # Suponha que as variáveis coluna_cetesb, coluna_epa e coluna_listaholandesa já foram definidas anteriormente

        # Encontrar o menor valor entre as variáveis
        menor_valor_entre_colunas = min(
            (coluna_cetesb, 'coluna_cetesb'),
            (coluna_epa, 'coluna_epa'),
            (coluna_listaholandesa, 'coluna_listaholandesa'),
            (coluna_conama, 'coluna_conama'),
            key=lambda x: x[0] if isinstance(x[0], int) else float('inf')
        )

        # Desempacotar o resultado para obter o valor e a variável correspondente
        menor_valor, variavel_menor_valor = menor_valor_entre_colunas

        # Encontrar o maior valor entre as variáveis
        maior_valor_entre_colunas = max(
            (coluna_cetesb, 'coluna_cetesb'),
            (coluna_epa, 'coluna_epa'),
            (coluna_listaholandesa, 'coluna_listaholandesa'),
            (coluna_conama, 'coluna_conama'),
            key=lambda x: x[0] if isinstance(x[0], int) else float('-inf')
        )

        # Desempacotar o resultado para obter o valor e a variável correspondente
        maior_valor, variavel_maior_valor = maior_valor_entre_colunas

        # Copiar os valores para as colunas D e E
        for row_idx in range(2, sheet_resultado.max_row + 1):
                    maior_valor_entre_colunas= sheet_resultado.cell(row=row_idx, column=maior_valor).value
                    menor_valor_entre_colunas = sheet_resultado.cell(row=row_idx, column=menor_valor).value

                    sheet_resultado.cell(row=row_idx, column=4, value=menor_valor_entre_colunas)
                    sheet_resultado.cell(row=row_idx, column=5, value=maior_valor_entre_colunas)

        for row in sheet_resultado.iter_rows(min_row=1, max_row=sheet_resultado.max_row, min_col=menor_valor, max_col=maior_valor):
            for cell in row:
                cell.value = None

        # Inserir o valor "Valor Orientador" na célula D1
        sheet_resultado.cell(row=1, column=4, value="Valor Orientador")

        # Redimensionar automaticamente largura e altura das células
        for col in sheet_resultado.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 1)
            sheet_resultado.column_dimensions[column].width = adjusted_width

        # Alinhar o texto ao centro das células mescladas e aplicar negrito
        for row in sheet_resultado.iter_rows(min_row=1, max_row=sheet_resultado.max_row, min_col=1, max_col=sheet_resultado.max_column - 2):
            for cell in row:

                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Aplicar negrito aos valores "Valor EPA" e "Valor Cetesb"
                if cell.value in ["Valor EPA", "Valor Cetesb","Valor Lista Holandesa","Valor Conama","Valor Orientador"]:
                    cell.font = Font(bold=True)

                # Adicionar todas as bordas com estilo "todas as bordas"
                cell.border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )

        # Mesclar a célula D1 até a célula E1
        sheet_resultado.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)

        # Mesclar células A1 até A2
        sheet_resultado.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        # Mesclar células B1 até B2
        sheet_resultado.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        # Mesclar células C1 até C3
        sheet_resultado.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

        # Verificação e coloração do texto das células a partir da coluna G e linha 3
        for row_idx in range(3, sheet_resultado.max_row + 1):
            for col_idx in range(6, sheet_resultado.max_column + 1):
                cell_value = sheet_resultado.cell(row=row_idx, column=col_idx).value

                # Verifica se o valor na célula é "<" e altera a cor do texto para preto
                if isinstance(cell_value, str) and "<" in cell_value:
                    sheet_resultado.cell(row=row_idx, column=col_idx).font = Font(color="C0C0C0")

    # Salvar as alterações no arquivo
    wb_resultado.save(caminho_analise)

if __name__ == "__main__":
    main()